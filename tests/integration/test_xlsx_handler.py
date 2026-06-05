from pathlib import Path
import pytest

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if HAS_OPENPYXL:
    from data_access_layer import XlsxHandler


@pytest.fixture
def temp_dir(tmp_path):
    """Create a temporary directory for test files."""
    return tmp_path


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxHandlerFetchIntegration:
    """Integration tests for XlsxHandler.fetch() with real files."""

    def test_fetch_from_real_xlsx_file(self, temp_dir):
        """Fetch data from an actual XLSX file."""
        from openpyxl import Workbook

        # Create test data file
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        # Fetch and verify
        handler = XlsxHandler()
        result = handler.fetch(path=test_file, table="Sheet")
        assert result == [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]

    def test_fetch_empty_file(self, temp_dir):
        """Fetch from an XLSX file with header only."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])

        test_file = temp_dir / "empty.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(path=test_file, table="Sheet")
        assert result == []

    def test_fetch_from_named_sheet(self, temp_dir):
        """Fetch from a specific named sheet."""
        from openpyxl import Workbook

        # Create file with multiple sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Users"
        ws1.append(["name", "age"])
        ws1.append(["Alice", 30])

        ws2 = wb.create_sheet("Products")
        ws2.append(["product", "price"])
        ws2.append(["Widget", 9.99])

        test_file = temp_dir / "data.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()

        # Fetch from Users sheet
        users = handler.fetch(path=test_file, table="Users")
        assert users == [{"name": "Alice", "age": 30}]

        # Fetch from Products sheet
        products = handler.fetch(path=test_file, table="Products")
        assert products == [{"product": "Widget", "price": 9.99}]

    def test_fetch_with_column_selection(self, temp_dir):
        """Fetch with column allowlist - only specified columns included."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "email"])
        ws.append(["Alice", 30, "alice@test.com"])
        ws.append(["Bob", 25, "bob@test.com"])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(path=test_file, table="Sheet", cols=["name", "age"])
        assert result == [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]

    def test_fetch_with_filter(self, temp_dir):
        """Filter is applied to rows."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(path=test_file, table="Sheet", filter_=lambda row: row["age"] > 25)
        assert result == [{"name": "Alice", "age": 30}]

    def test_fetch_with_limit(self, temp_dir):
        """Limit is applied after filtering."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(path=test_file, table="Sheet", limit=1)
        assert len(result) == 1
        assert result[0] == {"name": "Alice", "age": 30}

    def test_fetch_with_filter_and_limit(self, temp_dir):
        """Filter is applied first, then limit to filtered results."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        ws.append(["Charlie", 35])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(
            path=test_file, table="Sheet", filter_=lambda row: row["age"] >= 30, limit=1
        )
        # Two rows match filter (Alice 30, Charlie 35), but only 1 returned due to limit
        assert len(result) == 1

    def test_fetch_with_header_row_not_found_raises_value_error(self, temp_dir):
        """Fetching when header row is out of range raises ValueError."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler(header_row=5)
        with pytest.raises(ValueError, match="Header row 5 not found"):
            handler.fetch(path=test_file, table="Sheet", strict=True)

    def test_fetch_with_header_row_not_found_strict_false(self, temp_dir):
        """Fetching when header row is out of range returns empty list in lenient mode."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])

        test_file = temp_dir / "users.xlsx"
        wb.save(test_file)

        handler = XlsxHandler(header_row=5)
        result = handler.fetch(path=test_file, table="Sheet", strict=False)
        assert result == []

    def test_fetch_file_not_found(self, temp_dir):
        """Fetching when file doesn't exist raises FileNotFoundError."""
        handler = XlsxHandler()
        nonexistent_file = temp_dir / "missing.xlsx"
        with pytest.raises(FileNotFoundError, match="does not exist"):
            handler.fetch(path=nonexistent_file, table="Sheet", strict=True)

    def test_fetch_file_not_found_strict_false(self, temp_dir):
        """Fetching when file doesn't exist returns empty list in lenient mode."""
        handler = XlsxHandler()
        nonexistent_file = temp_dir / "missing.xlsx"
        result = handler.fetch(path=nonexistent_file, table="Sheet", strict=False)
        assert result == []


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxHandlerStoreIntegration:
    """Integration tests for XlsxHandler.store() with real files."""

    def test_store_to_real_xlsx_file(self, temp_dir):
        """Store data to an actual XLSX file."""
        from openpyxl import load_workbook

        test_data = [{"name": "Alice", "age": 30}]
        handler = XlsxHandler()

        test_file = temp_dir / "users.xlsx"
        result = handler.store(data=test_data, path=test_file, table="Sheet")
        assert result == 1

        # Verify file was written correctly
        wb = load_workbook(test_file)
        ws = wb.active
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30

    def test_store_to_named_sheet(self, temp_dir):
        """Store to a specific named sheet."""
        from openpyxl import load_workbook

        test_data = [{"name": "Alice", "age": 30}]
        handler = XlsxHandler()

        test_file = temp_dir / "data.xlsx"
        handler.store(data=test_data, path=test_file, table="Users")

        # Verify sheet name
        wb = load_workbook(test_file)
        assert "Users" in wb.sheetnames

    def test_store_overwrite_replaces_sheet_content(self, temp_dir):
        """Store with overwrite=True replaces existing sheet content."""
        from openpyxl import Workbook, load_workbook

        handler = XlsxHandler()

        # Create initial data
        test_file = temp_dir / "data.xlsx"
        initial_data = [{"name": "Alice", "age": 30}]
        handler.store(path=test_file, table="Users", data=initial_data)

        # Overwrite with new data
        new_data = [{"name": "Bob", "age": 25}]
        handler.store(path=test_file, table="Users", data=new_data, overwrite=True)

        # Verify only new data exists
        wb = load_workbook(test_file)
        ws = wb["Users"]
        assert ws["A2"].value == "Bob"
        assert ws["B2"].value == 25
        # Should only have header + 1 row
        assert ws.max_row == 2

    def test_store_append_adds_to_existing_sheet(self, temp_dir):
        """Store with overwrite=False appends to existing sheet."""
        from openpyxl import load_workbook

        handler = XlsxHandler()

        # Create initial data
        test_file = temp_dir / "data.xlsx"
        initial_data = [{"name": "Alice", "age": 30}]
        handler.store(path=test_file, table="Users", data=initial_data)

        # Append more data
        more_data = [{"name": "Bob", "age": 25}]
        handler.store(path=test_file, table="Users", data=more_data, overwrite=False)

        # Verify both rows exist
        wb = load_workbook(test_file)
        ws = wb["Users"]
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        assert ws["A3"].value == "Bob"
        assert ws["B3"].value == 25
        # Should have header + 2 rows
        assert ws.max_row == 3

    def test_store_with_column_selection(self, temp_dir):
        """Store with column allowlist - only specified columns stored."""
        from openpyxl import load_workbook

        test_data = [{"name": "Alice", "age": 30, "email": "alice@test.com"}]
        handler = XlsxHandler()

        test_file = temp_dir / "users.xlsx"
        handler.store(data=test_data, path=test_file, table="Sheet", cols=["name", "age"])

        # Verify only specified columns stored
        wb = load_workbook(test_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("name", "age")
        assert rows[1] == ("Alice", 30)
        assert len(rows) == 2

    def test_store_with_filter(self, temp_dir):
        """Filter is applied before storing."""
        from openpyxl import load_workbook

        test_data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25}
        ]
        handler = XlsxHandler()

        test_file = temp_dir / "users.xlsx"
        handler.store(data=test_data, path=test_file, table="Sheet", filter_=lambda row: row["age"] > 25)

        # Verify only filtered data stored
        wb = load_workbook(test_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("name", "age")
        assert rows[1] == ("Alice", 30)
        assert len(rows) == 2

    def test_store_with_limit(self, temp_dir):
        """Limit is applied after filtering."""
        from openpyxl import load_workbook

        test_data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25}
        ]
        handler = XlsxHandler()

        test_file = temp_dir / "users.xlsx"
        handler.store(data=test_data, path=test_file, table="Sheet", limit=1)

        # Verify only 1 row stored
        wb = load_workbook(test_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("name", "age")
        assert rows[1] == ("Alice", 30)
        assert len(rows) == 2

    def test_store_raises_error_when_parent_directory_does_not_exist(self, temp_dir):
        """Store should raise error when parent directory doesn't exist."""
        handler = XlsxHandler()
        data = [{"name": "Alice", "age": 30}]

        # Path with non-existent parent directory
        nonexistent_file = temp_dir / "subdir" / "data.xlsx"

        with pytest.raises(FileNotFoundError, match="does not exist"):
            handler.store(path=nonexistent_file, table="Users", data=data)


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxHandlerTypeCoercion:
    """Integration tests for XlsxHandler type coercion functionality."""

    def test_fetch_with_types_coerces_values(self, temp_dir):
        """Type coercion in fetch() converts string values to specified types."""
        from openpyxl import Workbook

        # Create test data with string values that should be coerced
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "name", "age"])
        ws.append(["1", "Alice", "30"])
        ws.append(["2", "Bob", "25"])

        test_file = temp_dir / "test_types.xlsx"
        wb.save(test_file)

        # Fetch with type coercion
        handler = XlsxHandler()
        result = handler.fetch(
            path=test_file,
            table="Sheet",
            types={'id': int, 'age': int}
        )

        # Verify values were coerced to correct types
        assert result[0]['id'] == 1
        assert isinstance(result[0]['id'], int)
        assert result[0]['age'] == 30
        assert isinstance(result[0]['age'], int)
        assert result[0]['name'] == "Alice"
        assert isinstance(result[0]['name'], str)

        assert result[1]['id'] == 2
        assert isinstance(result[1]['id'], int)
        assert result[1]['age'] == 25
        assert isinstance(result[1]['age'], int)

    def test_fetch_with_partial_type_coercion(self, temp_dir):
        """Type coercion only applies to specified columns."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "name", "age", "score"])
        ws.append(["1", "Alice", "30", "95.5"])

        test_file = temp_dir / "partial_types.xlsx"
        wb.save(test_file)

        handler = XlsxHandler()
        result = handler.fetch(
            path=test_file,
            table="Sheet",
            types={'id': int}  # Only coerce id
        )

        # id should be coerced
        assert result[0]['id'] == 1
        assert isinstance(result[0]['id'], int)

        # Other columns remain as-is
        assert result[0]['name'] == "Alice"
        assert result[0]['age'] == "30"
        assert result[0]['score'] == "95.5"

    def test_store_with_types_coerces_values(self, temp_dir):
        """Type coercion in store() converts values before writing."""
        from openpyxl import load_workbook

        test_data = [
            {'id': '1', 'name': 'Alice', 'age': '30'},
            {'id': '2', 'name': 'Bob', 'age': '25'}
        ]

        handler = XlsxHandler()
        test_file = temp_dir / "store_types.xlsx"

        # Store with type coercion
        handler.store(
            data=test_data,
            path=test_file,
            table="Sheet",
            types={'id': int, 'age': int}
        )

        # Verify values were coerced and stored correctly
        wb = load_workbook(test_file)
        ws = wb.active

        # Check header
        assert ws["A1"].value == "id"
        assert ws["B1"].value == "name"
        assert ws["C1"].value == "age"

        # Check first data row - values should be coerced types
        assert ws["A2"].value == 1
        assert isinstance(ws["A2"].value, int)
        assert ws["B2"].value == "Alice"
        assert ws["C2"].value == 30
        assert isinstance(ws["C2"].value, int)

        # Check second data row
        assert ws["A3"].value == 2
        assert isinstance(ws["A3"].value, int)
        assert ws["B3"].value == "Bob"
        assert ws["C3"].value == 25
        assert isinstance(ws["C3"].value, int)

    def test_store_and_fetch_with_types_roundtrip(self, temp_dir):
        """Store and fetch with type coercion maintains data integrity."""
        test_data = [
            {'id': '1', 'name': 'Alice', 'score': '100'},
            {'id': '2', 'name': 'Bob', 'score': '95'}
        ]

        handler = XlsxHandler()
        test_file = temp_dir / "roundtrip.xlsx"

        # Store with type coercion
        handler.store(
            data=test_data,
            path=test_file,
            table="Data",
            types={'id': int, 'score': int}
        )

        # Fetch with same type coercion
        result = handler.fetch(
            path=test_file,
            table="Data",
            types={'id': int, 'score': int}
        )

        # Verify data integrity
        assert len(result) == 2
        assert result[0] == {'id': 1, 'name': 'Alice', 'score': 100}
        assert result[1] == {'id': 2, 'name': 'Bob', 'score': 95}

        # Verify types
        assert isinstance(result[0]['id'], int)
        assert isinstance(result[0]['score'], int)
