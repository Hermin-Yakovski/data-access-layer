import pytest
from pathlib import Path
from data_access_layer.xlsx_handler import AsyncXlsxHandler

@pytest.mark.asyncio
async def test_async_xlsx_fetch_basic(tmp_path):
    """Test basic async fetch from XLSX file."""
    from openpyxl import Workbook
    test_file = tmp_path / "test.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["name", "age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    wb.save(test_file)

    handler = AsyncXlsxHandler()
    result = await handler.fetch(test_file, "Sheet")

    assert len(result) == 2
    assert result[0] == {"name": "Alice", "age": 30}

@pytest.mark.asyncio
async def test_async_xlsx_store_basic(tmp_path):
    """Test basic async store to XLSX file."""
    handler = AsyncXlsxHandler()
    test_data = [{"name": "Charlie", "age": 35}]
    test_file = tmp_path / "output.xlsx"
    rows_written = await handler.store(test_data, test_file, "Sheet")

    assert rows_written == 1

    # Verify file exists and can be read
    from openpyxl import load_workbook
    wb = load_workbook(test_file)
    ws = wb.active
    assert ws["A1"].value == "name"
    assert ws["A2"].value == "Charlie"
