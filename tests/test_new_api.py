# tests/test_new_api.py
"""Tests for the new XlsxHandler API.

This file tests the redesigned XlsxHandler API where:
- __init__ only takes header_row parameter (not sheet_name)
- sheet_name is specified via the table parameter in fetch() and store()
"""
from pathlib import Path
import pytest

try:
    from dal.xlsx_handler import XlsxHandler
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestNewAPI:
    """Tests for the redesigned XlsxHandler API."""

    def test_init_accepts_only_header_row(self):
        """XlsxHandler.__init__ should accept only header_row parameter."""
        # Should work with default header_row
        handler = XlsxHandler()
        assert hasattr(handler, 'header_row')
        assert handler.header_row == 0

        # Should work with explicit header_row
        handler = XlsxHandler(header_row=1)
        assert handler.header_row == 1

    def test_init_does_not_accept_sheet_name(self):
        """XlsxHandler.__init__ should not accept sheet_name parameter."""
        # This should raise TypeError due to unexpected keyword argument
        with pytest.raises(TypeError, match="sheet_name"):
            handler = XlsxHandler(sheet_name="MySheet")

    def test_init_does_not_have_sheet_name_attribute(self):
        """XlsxHandler instance should not have sheet_name attribute."""
        handler = XlsxHandler()
        assert not hasattr(handler, 'sheet_name')

    def test_fetch_uses_path_as_file_path_and_table_as_sheet_name(self, tmp_path):
        """fetch() should use path as file path and table as sheet name."""
        from openpyxl import Workbook

        # Create a test file with multiple sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Users"
        ws1.append(["name", "age"])
        ws1.append(["Alice", 30])

        ws2 = wb.create_sheet("Products")
        ws2.append(["product", "price"])
        ws2.append(["Widget", 9.99])

        # Save to a specific file path
        test_file = tmp_path / "data.xlsx"
        wb.save(test_file)

        # Test fetching with path as file and table as sheet
        handler = XlsxHandler()

        # Fetch from Users sheet
        users = handler.fetch(path=test_file, table="Users")
        assert users == [{"name": "Alice", "age": 30}]

        # Fetch from Products sheet
        products = handler.fetch(path=test_file, table="Products")
        assert products == [{"product": "Widget", "price": 9.99}]

    def test_fetch_raises_error_when_file_not_found(self, tmp_path):
        """fetch() should raise FileNotFoundError when file path doesn't exist."""
        handler = XlsxHandler()
        nonexistent_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError, match="does not exist"):
            handler.fetch(path=nonexistent_file, table="Sheet1")

    def test_store_uses_path_as_file_path_and_table_as_sheet_name(self, tmp_path):
        """store() should use path as file path and table as sheet name."""
        # Test data
        users_data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25}
        ]

        handler = XlsxHandler()

        # Store data to a specific file path with a specific sheet name
        test_file = tmp_path / "data.xlsx"
        rows_stored = handler.store(path=test_file, table="Users", data=users_data)

        assert rows_stored == 2

        # Verify the file was created at the correct location
        assert test_file.exists()

        # Verify the data was stored in the correct sheet
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        assert "Users" in wb.sheetnames

        # Verify the data in the Users sheet
        ws = wb["Users"]
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        assert ws["A3"].value == "Bob"
        assert ws["B3"].value == 25

    def test_store_overwrite_mode_replaces_sheet_content(self, tmp_path):
        """store() with overwrite=True should replace existing sheet content."""
        from openpyxl import Workbook, load_workbook

        handler = XlsxHandler()

        # Create initial data
        test_file = tmp_path / "data.xlsx"
        initial_data = [{"name": "Alice", "age": 30}]
        handler.store(path=test_file, table="Users", data=initial_data)

        # Overwrite with new data
        new_data = [{"name": "Bob", "age": 25}]
        handler.store(path=test_file, table="Users", data=new_data, overwrite=True)

        # Verify only new data exists
        wb = load_workbook(test_file)
        ws = wb["Users"]
        assert ws["A2"].value == "Bob"
        assert ws["B2"].value == 25
        # Should only have header + 1 row
        assert ws.max_row == 2

    def test_store_append_mode_adds_to_existing_sheet(self, tmp_path):
        """store() with overwrite=False should append to existing sheet."""
        from openpyxl import load_workbook

        handler = XlsxHandler()

        # Create initial data
        test_file = tmp_path / "data.xlsx"
        initial_data = [{"name": "Alice", "age": 30}]
        handler.store(path=test_file, table="Users", data=initial_data)

        # Append more data
        more_data = [{"name": "Bob", "age": 25}]
        handler.store(path=test_file, table="Users", data=more_data, overwrite=False)

        # Verify both rows exist
        wb = load_workbook(test_file)
        ws = wb["Users"]
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        assert ws["A3"].value == "Bob"
        assert ws["B3"].value == 25
        # Should have header + 2 rows
        assert ws.max_row == 3

    def test_store_raises_error_when_parent_directory_does_not_exist(self, tmp_path):
        """store() should raise error when parent directory doesn't exist."""
        handler = XlsxHandler()
        data = [{"name": "Alice", "age": 30}]

        # Path with non-existent parent directory
        nonexistent_file = tmp_path / "subdir" / "data.xlsx"

        with pytest.raises(FileNotFoundError, match="does not exist"):
            handler.store(path=nonexistent_file, table="Users", data=data)
