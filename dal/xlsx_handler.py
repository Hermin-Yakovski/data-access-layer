import asyncio
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Type

try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .abc import AsyncDataHandler, DataHandler
from .post_processing import PostProcessingMixin


class XlsxHandler(PostProcessingMixin, DataHandler):
    """Handler for Excel XLSX format files.

    Supports fetching and storing data in XLSX format with optional
    header_row configuration.

    Requires openpyxl to be installed. If not available, imports will fail.
    """

    def __init__(self, header_row: int = 0):
        """Initialize XlsxHandler.

        Args:
            header_row: Row number containing column headers (0-indexed, default: 0)

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for XlsxHandler. Install it with: pip install openpyxl"
            )
        self.header_row = header_row

    def fetch(
        self,
        path: Path,
        table: str,
        cols: Optional[Iterable[str]] = None,
        filter_: Optional[Callable[[Dict[str, Any]], bool]] = None,
        limit: Optional[int] = None,
        types: Optional[Dict[str, Type]] = None,
        strict: bool = True,
    ) -> List[Dict[str, Any]]:
        """Fetch data from XLSX file.

        Args:
            path: Path to the XLSX file
            table: Sheet name to fetch from
            cols: Columns to include (allowlist, None = all columns)
            filter_: Optional callable for row filtering
            limit: Maximum rows to return (applied after filtering)
            types: Optional dict mapping field names to target types for coercion
            strict: If True, raise exceptions; if False, return empty list on error

        Returns:
            List of row dictionaries
        """
        try:
            if not path.exists():
                raise FileNotFoundError(f"File '{path}' does not exist")

            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb[table]

                # Get header row
                headers = None
                rows_data = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == self.header_row:
                        headers = [
                            str(cell) if cell is not None else f"col_{i}"
                            for i, cell in enumerate(row)
                        ]
                    elif row_idx > self.header_row:
                        rows_data.append(row)

                if headers is None:
                    raise ValueError(f"Header row {self.header_row} not found in sheet")

                # Convert rows to dictionaries
                data = []
                for row in rows_data:
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                    data.append(row_dict)
            finally:
                wb.close()

            # Apply post-processing
            data = self._apply_processing(data, types, cols, filter_, limit)

            return data

        except Exception:
            if strict:
                raise
            return []

    def store(
        self,
        data: List[Dict[str, Any]],
        path: Path,
        table: str,
        cols: Optional[Iterable[str]] = None,
        filter_: Optional[Callable[[Dict[str, Any]], bool]] = None,
        limit: Optional[int] = None,
        types: Optional[Dict[str, Type]] = None,
        overwrite: bool = True,
        strict: bool = True,
    ) -> int:
        """Store data to XLSX file.

        Args:
            data: List of row dictionaries to store
            path: Path to the XLSX file
            table: Sheet name to store to
            cols: Columns to include (allowlist, None = all columns)
            filter_: Optional callable for row filtering
            limit: Maximum rows to store (applied after filtering)
            types: Optional dict mapping field names to target types for coercion
            overwrite: If True, replace existing sheet content; if False, append
            strict: If True, raise exceptions; if False, return 0 on error

        Returns:
            Number of rows stored
        """
        try:
            # Check if parent directory exists
            if path.parent != Path('.') and not path.parent.exists():
                raise FileNotFoundError(f"Parent directory '{path.parent}' does not exist")

            # Prepare data to store
            data_to_store = data.copy()

            # Apply post-processing
            data_to_store = self._apply_processing(data_to_store, types, cols, filter_, limit)

            # Handle workbook creation or loading
            if path.exists():
                wb = load_workbook(path)
            else:
                wb = Workbook()
                # Remove the default sheet if we're going to create a named one
                if wb.active.title == "Sheet" and table != "Sheet":
                    wb.remove(wb.active)

            # For append mode, read existing data and merge
            will_merge_data = not overwrite and path.exists() and table in wb.sheetnames
            if will_merge_data:
                existing_data = self.fetch(path=path, table=table, strict=True)
                data_to_store = existing_data + data_to_store

            # Handle sheet creation or selection
            if table in wb.sheetnames:
                # Remove the sheet and create a new one (for both overwrite and merge modes)
                wb.remove(wb[table])
                ws = wb.create_sheet(title=table, index=0)
            else:
                ws = wb.create_sheet(title=table)

            # Set the sheet title
            ws.title = table

            if data_to_store:
                # Write header
                headers = list(data_to_store[0].keys())
                ws.append(headers)

                # Write data rows
                for row in data_to_store:
                    ws.append([row.get(h, "") for h in headers])

            # Save workbook
            wb.save(path)

            return len(data_to_store)

        except Exception:
            if strict:
                raise
            return 0


class AsyncXlsxHandler(PostProcessingMixin, AsyncDataHandler):
    """Async handler for Excel XLSX format files.

    Supports fetching and storing data in XLSX format with optional
    header_row configuration. Uses asyncio.to_thread for CPU-bound
    openpyxl operations.

    Requires openpyxl to be installed. If not available, imports will fail.
    """

    def __init__(self, header_row: int = 0):
        """Initialize AsyncXlsxHandler.

        Args:
            header_row: Row number containing column headers (0-indexed, default: 0)

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for AsyncXlsxHandler. Install it with: pip install openpyxl"
            )
        self.header_row = header_row

    async def fetch(
        self,
        path: Path,
        table: str,
        cols: Optional[Iterable[str]] = None,
        filter_: Optional[Callable[[Dict[str, Any]], bool]] = None,
        limit: Optional[int] = None,
        types: Optional[Dict[str, Type]] = None,
        strict: bool = True,
    ) -> List[Dict[str, Any]]:
        """Fetch data from XLSX file asynchronously.

        Args:
            path: Path to the XLSX file
            table: Sheet name to fetch from
            cols: Columns to include (allowlist, None = all columns)
            filter_: Optional callable for row filtering
            limit: Maximum rows to return (applied after filtering)
            types: Optional dict mapping field names to target types for coercion
            strict: If True, raise exceptions; if False, return empty list on error

        Returns:
            List of row dictionaries
        """
        try:
            if not path.exists():
                raise FileNotFoundError(f"File '{path}' does not exist")

            def _load_workbook():
                wb = load_workbook(path, read_only=True, data_only=True)
                try:
                    ws = wb[table]

                    # Get header row
                    headers = None
                    rows_data = []
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        if row_idx == self.header_row:
                            headers = [
                                str(cell) if cell is not None else f"col_{i}"
                                for i, cell in enumerate(row)
                            ]
                        elif row_idx > self.header_row:
                            rows_data.append(row)

                    if headers is None:
                        raise ValueError(f"Header row {self.header_row} not found in sheet")

                    # Convert rows to dictionaries
                    data = []
                    for row in rows_data:
                        row_dict = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                row_dict[header] = row[i]
                        data.append(row_dict)
                finally:
                    wb.close()

                return data

            data = await asyncio.to_thread(_load_workbook)

            # Apply post-processing
            data = self._apply_processing(data, types, cols, filter_, limit)

            return data

        except Exception:
            if strict:
                raise
            return []

    async def store(
        self,
        data: List[Dict[str, Any]],
        path: Path,
        table: str,
        cols: Optional[Iterable[str]] = None,
        filter_: Optional[Callable[[Dict[str, Any]], bool]] = None,
        limit: Optional[int] = None,
        types: Optional[Dict[str, Type]] = None,
        overwrite: bool = True,
        strict: bool = True,
    ) -> int:
        """Store data to XLSX file asynchronously.

        Args:
            data: List of row dictionaries to store
            path: Path to the XLSX file
            table: Sheet name to store to
            cols: Columns to include (allowlist, None = all columns)
            filter_: Optional callable for row filtering
            limit: Maximum rows to store (applied after filtering)
            types: Optional dict mapping field names to target types for coercion
            overwrite: If True, replace existing sheet content; if False, append
            strict: If True, raise exceptions; if False, return 0 on error

        Returns:
            Number of rows stored
        """
        try:
            # Check if parent directory exists
            if path.parent != Path('.') and not path.parent.exists():
                raise FileNotFoundError(f"Parent directory '{path.parent}' does not exist")

            # Prepare data to store
            data_to_store = data.copy()

            # Apply post-processing
            data_to_store = self._apply_processing(data_to_store, types, cols, filter_, limit)

            def _write_workbook(data_to_store):
                # Handle workbook creation or loading
                if path.exists():
                    wb = load_workbook(path)
                else:
                    wb = Workbook()
                    # Remove the default sheet if we're going to create a named one
                    if wb.active.title == "Sheet" and table != "Sheet":
                        wb.remove(wb.active)

                # For append mode, we need to read existing data synchronously
                will_merge_data = not overwrite and path.exists() and table in wb.sheetnames
                if will_merge_data:
                    # Use sync fetch for merge (we're already in a thread)
                    if table in wb.sheetnames:
                        ws = wb[table]
                        existing_data = []
                        headers = None
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if row_idx == self.header_row:
                                headers = list(row)
                            elif row_idx > self.header_row:
                                row_dict = {}
                                for i, header in enumerate(headers):
                                    if i < len(row):
                                        row_dict[header] = row[i]
                                existing_data.append(row_dict)
                        data_to_store = existing_data + data_to_store

                # Handle sheet creation or selection
                if table in wb.sheetnames:
                    # Remove the sheet and create a new one (for both overwrite and merge modes)
                    wb.remove(wb[table])
                    ws = wb.create_sheet(title=table, index=0)
                else:
                    ws = wb.create_sheet(title=table)

                # Set the sheet title
                ws.title = table

                if data_to_store:
                    # Write header
                    headers = list(data_to_store[0].keys())
                    ws.append(headers)

                    # Write data rows
                    for row in data_to_store:
                        ws.append([row.get(h, "") for h in headers])

                # Save workbook
                wb.save(path)
                wb.close()

                return len(data_to_store)

            return await asyncio.to_thread(_write_workbook, data_to_store)

        except Exception:
            if strict:
                raise
            return 0
